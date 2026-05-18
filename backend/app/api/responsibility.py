"""
责任田 API
支持：Excel 上传、解析和展示
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import Optional, List, Any
import json
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

from ..database import get_db
from ..models.user import User, UserRole
from ..services.auth import get_current_user, get_current_admin

router = APIRouter()

# 责任田数据存储路径
RESPONSIBILITY_DATA_FILE = "/app/uploads/responsibility_field.json"


def ensure_upload_dir():
    """确保上传目录存在"""
    os.makedirs("/app/uploads", exist_ok=True)


def load_responsibility_data() -> Optional[dict]:
    """加载责任田数据"""
    try:
        if os.path.exists(RESPONSIBILITY_DATA_FILE):
            with open(RESPONSIBILITY_DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return None


def save_responsibility_data(data: dict):
    """保存责任田数据"""
    ensure_upload_dir()
    with open(RESPONSIBILITY_DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def parse_excel(file_content: bytes) -> dict:
    """
    解析 Excel 文件为 JSON 格式
    返回包含所有 sheet 的数据
    """
    wb = load_workbook(filename=BytesIO(file_content), data_only=True)
    result = {
        "sheets": [],
        "uploadedAt": datetime.now().isoformat()
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 获取合并单元格信息
        merged_cells = []
        for merged_range in ws.merged_cells.ranges:
            merged_cells.append({
                "startRow": merged_range.min_row - 1,  # 转为0索引
                "endRow": merged_range.max_row - 1,
                "startCol": merged_range.min_col - 1,
                "endCol": merged_range.max_col - 1
            })

        # 获取列宽信息
        col_widths = []
        max_col = ws.max_column if ws.max_column else 1
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            col_widths.append(width if width else 10)  # 默认宽度

        # 读取数据
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            row_data = []
            for cell in row:
                value = cell.value
                if value is None:
                    value = ""
                elif isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d")
                else:
                    value = str(value)
                row_data.append(value)
            rows.append(row_data)

        # 移除末尾的空行
        while rows and all(cell == "" for cell in rows[-1]):
            rows.pop()

        sheet_data = {
            "name": sheet_name,
            "rows": rows,
            "colWidths": col_widths,
            "mergedCells": merged_cells,
            "maxRow": len(rows),
            "maxCol": ws.max_column
        }
        result["sheets"].append(sheet_data)

    return result


@router.get("")
def get_responsibility_field(
    current_user: User = Depends(get_current_user)
):
    """
    获取责任田数据
    所有登录用户都可以查看
    """
    data = load_responsibility_data()
    if data:
        return {"success": True, "data": data}
    return {"success": True, "data": None}


@router.post("/upload")
async def upload_responsibility_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_admin)  # 只有 admin 可以上传
):
    """
    上传责任田 Excel 文件
    仅管理员可操作
    """
    # 验证文件类型
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名不能为空")

    allowed_extensions = ['.xlsx', '.xls']
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

    try:
        # 读取文件内容
        content = await file.read()

        # 解析 Excel
        data = parse_excel(content)
        data["originalFilename"] = file.filename
        data["uploadedBy"] = current_user.name
        data["uploadedById"] = current_user.id

        # 保存数据
        save_responsibility_data(data)

        return {
            "success": True,
            "message": "责任田数据已更新",
            "data": data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"解析 Excel 失败: {str(e)}")


@router.delete("")
def delete_responsibility_field(
    current_user: User = Depends(get_current_admin)  # 只有 admin 可以删除
):
    """
    删除责任田数据
    仅管理员可操作
    """
    try:
        if os.path.exists(RESPONSIBILITY_DATA_FILE):
            os.remove(RESPONSIBILITY_DATA_FILE)
        return {"success": True, "message": "责任田数据已删除"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


@router.get("/fields")
def get_responsibility_fields(
    current_user: User = Depends(get_current_user)
):
    """
    从 Excel 数据中提取结构化的责任田字段列表
    供 Capability 页面选择责任田归属
    """
    data = load_responsibility_data()
    if not data:
        return {"success": True, "fields": []}

    sheets = data.get("sheets", [])
    if not sheets:
        return {"success": True, "fields": []}

    # Take the first sheet
    sheet = sheets[0]
    rows = sheet.get("rows", [])

    if len(rows) < 2:
        return {"success": True, "fields": []}

    # First row is header
    headers = rows[0]
    data_rows = rows[1:]

    # Infer column indices
    def find_col_index(keywords):
        for i, h in enumerate(headers):
            h_lower = h.strip().lower()
            for kw in keywords:
                if kw in h_lower:
                    return i
        return None

    name_col = find_col_index(["责任田", "领域", "field", "name", "名称", "模块"]) or 0
    owner_col = find_col_index(["负责人", "owner", "人", "dri", "主责"]) or 1 if len(headers) > 1 else None

    fields = []
    for row_idx, row in enumerate(data_rows):
        if not row or all(c == "" for c in row):
            continue

        name = row[name_col].strip() if name_col < len(row) and row[name_col] else ""
        if not name:
            continue

        owner_name = ""
        if owner_col is not None and owner_col < len(row):
            owner_name = row[owner_col].strip()

        # Build a key-value map of all columns
        raw = {}
        for col_idx, val in enumerate(row):
            if col_idx < len(headers):
                raw[headers[col_idx]] = val if val else ""
            else:
                raw[f"col_{col_idx}"] = val if val else ""

        fields.append({
            "id": row_idx + 1,
            "name": name,
            "ownerName": owner_name,
            "raw": raw,
        })

    return {"success": True, "fields": fields}
